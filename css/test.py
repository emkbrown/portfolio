import os
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("MySheet1", -1)
ws2 = wb.create_sheet("MySheet4")
print(wb.sheetnames)